import streamlit as st
import pandas as pd
import csv
import os
import openpyxl
from openpyxl.styles import PatternFill
from PIL import Image
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from thefuzz import fuzz  # Import para fuzzy matching
import sys # Importado para ajustar o limite do campo CSV

CONFIG_FILE = 'configuracoes.csv'

st.set_page_config(page_title="Sistema de Auditoria Tributária - Escritório Contábil Sigilo", layout="centered")

# --- Correção do Limite de Campo CSV ---
# Aumenta o limite do tamanho do campo para evitar erros com células grandes no CSV.
# Usar sys.maxsize pode consumir muita memória se um campo for extremamente grande.
# Um limite grande e fixo pode ser uma alternativa, mas sys.maxsize é a abordagem comum.
try:
    max_int = sys.maxsize
    decrement = True
    while decrement:
        # Diminui o max_int até que funcione
        try:
            csv.field_size_limit(max_int)
            decrement = False
        except OverflowError:
            max_int = int(max_int / 10)
except Exception as e:
    st.warning(f"Não foi possível ajustar o limite do campo CSV dinamicamente: {e}. Usando limite padrão.")

# Carregar logo (se presente)
try:
    logo = Image.open("logo.png")
    st.image(logo, width=250)
except FileNotFoundError:
    st.warning("⚠️ A logo 'logo.png' não foi encontrada. Coloque na mesma pasta do app.py.")

st.title("📊 Sistema de Auditoria Tributária de Produtos - Escritório Contábil Sigilo")

def clean_cest(cest_value):
    """Limpa o valor do CEST, removendo '.0' e garantindo que seja uma string."""
    if pd.isna(cest_value):
        return '0'
    cest_str = str(cest_value).strip()
    if cest_str.endswith('.0'):
        cest_str = cest_str[:-2] # Remove '.0'
    # Tenta converter para int para remover quaisquer outros decimais e depois para str
    try:
        return str(int(cest_str))
    except ValueError:
        # Se não for um número válido após remover .0, retorna a string limpa
        return cest_str if cest_str else '0'

def get_keywords(text):
    stopwords = {'de', 'da', 'do', 'e', 'em', 'com', 'ml'}
    return set(word.lower() for word in str(text).split() if word.lower() not in stopwords and len(word) > 2)

def load_configurations():
    configs = {}
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                try:
                    sample = f.read(2048) # Ler uma amostra maior pode ajudar
                    dialect = csv.Sniffer().sniff(sample)
                    f.seek(0)
                except csv.Error:
                    f.seek(0)
                    dialect = csv.excel # Usa um dialeto padrão
                    st.sidebar.warning("Não foi possível detectar o delimitador do CSV automaticamente, tentando com vírgula.")

                reader = csv.reader(f, dialect)
                header = next(reader, None) # Pula o cabeçalho se existir

                for row_num, row in enumerate(reader, start=2): # start=2 para contar cabeçalho
                    try:
                        if len(row) >= 5: # Espera 5 colunas: Desc, NCM, ALIQ, TRIB, CEST
                            desc_val = str(row[0]).strip()
                            ncm_val = str(row[1]).strip()
                            aliq_val = str(row[2]).strip()
                            trib_val = str(row[3]).strip() # Nome interno sem Ç
                            cest_val = clean_cest(row[4])
                            configs[desc_val] = {
                                'NCM': ncm_val,
                                'ALIQ_ICMS': aliq_val,
                                'TRIBUTACAO': trib_val, # Padronizado sem Ç
                                'CEST': cest_val
                            }
                        elif len(row) == 4: # Caso antigo sem CEST explícito
                             desc_val = str(row[0]).strip()
                             ncm_val = str(row[1]).strip()
                             aliq_val = str(row[2]).strip()
                             trib_val = str(row[3]).strip()
                             configs[desc_val] = {
                                'NCM': ncm_val,
                                'ALIQ_ICMS': aliq_val,
                                'TRIBUTACAO': trib_val, # Padronizado sem Ç
                                'CEST': '0'
                            }
                        # Ignora linhas com menos de 4 colunas silenciosamente ou adiciona um aviso
                        # else:
                        #     st.sidebar.warning(f"Linha {row_num} no CSV ignorada: número inesperado de colunas ({len(row)}).")
                    except IndexError as ie:
                         st.sidebar.error(f"Erro de índice ao processar linha {row_num} do CSV: {ie}. Verifique a estrutura do arquivo.")
                         continue # Pula para a próxima linha

            if configs:
                 st.sidebar.success(f"✅ Arquivo CSV '{CONFIG_FILE}' carregado com sucesso! {len(configs)} itens encontrados.")
            else:
                 st.sidebar.warning(f"Arquivo CSV '{CONFIG_FILE}' carregado, mas nenhum item válido encontrado.")

        except Exception as e:
            # Verifica se o erro é o de limite de campo, mesmo com o ajuste (pode indicar outros problemas)
            if 'field larger than field limit' in str(e):
                 st.sidebar.error(f"Erro ao carregar CSV '{CONFIG_FILE}': O arquivo contém um campo excessivamente grande, mesmo após tentar ajustar o limite. Verifique o arquivo. Detalhe: {str(e)}")
            else:
                 st.sidebar.error(f"Erro ao carregar arquivo CSV '{CONFIG_FILE}': {str(e)}")

    # Se o CSV estava vazio ou não existe, tenta carregar do Excel
    if not configs and os.path.exists('configuracoes.xlsx'):
        st.sidebar.info("Arquivo CSV não encontrado ou vazio. Tentando carregar de 'configuracoes.xlsx'...")
        try:
            df = pd.read_excel('configuracoes.xlsx', dtype={'NCM': str, 'CEST': str})
            # Limpeza robusta de nomes de coluna
            df.columns = df.columns.str.strip().str.replace('"', '', regex=False).str.replace('\n', '', regex=False).str.replace('\r', '', regex=False).str.upper()
            df.columns = df.columns.str.replace('.', '', regex=False) # Remove pontos comuns em nomes de coluna

            # Mapeamento flexível de nomes de coluna (ajuste conforme variações comuns)
            col_mapping = {
                'DESCRIÇÃO ITEM': 'Descrição item', 'DESCRICAO ITEM': 'Descrição item',
                'NCM': 'NCM',
                'ALIQ ICMS': 'Aliq. ICMS', 'ALIQUOTA ICMS': 'Aliq. ICMS',
                'TRIBUTAÇÃO': 'TRIBUTACAO', 'TRIBUTACAO': 'TRIBUTACAO',
                'CEST': 'CEST'
            }
            df.rename(columns=col_mapping, inplace=True)

            required_cols = ['Descrição item', 'NCM', 'Aliq. ICMS']
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                st.sidebar.error(f"Erro: O arquivo Excel deve conter as colunas: {', '.join(required_cols)}. Colunas ausentes ou não reconhecidas: {', '.join(missing_cols)}")
                return {}

            for _, row in df.iterrows():
                desc = str(row['Descrição item']).strip()
                if not desc: continue

                ncm = str(row['NCM']).strip() if pd.notna(row['NCM']) else ''
                aliq = str(row['Aliq. ICMS']).strip()
                # Usa .get com o nome padronizado sem Ç e fallback para Aliq. ICMS
                trib = str(row.get('TRIBUTACAO', aliq)).strip()
                cest = clean_cest(row.get('CEST', '0'))

                configs[desc] = {'NCM': ncm, 'ALIQ_ICMS': aliq, 'TRIBUTACAO': trib, 'CEST': cest}

            if configs:
                save_all_configurations(configs) # Salva no formato CSV padrão
                st.sidebar.success(f"✅ Base carregada do arquivo Excel 'configuracoes.xlsx' e salva em '{CONFIG_FILE}'. {len(configs)} itens encontrados.")
            else:
                 st.sidebar.warning("Nenhum item válido encontrado no arquivo Excel 'configuracoes.xlsx'.")

        except Exception as e:
            st.sidebar.error(f"Erro ao carregar arquivo Excel 'configuracoes.xlsx': {str(e)}")

    return configs

def save_all_configurations(configs):
    """Salva todas as configurações no arquivo CSV, garantindo que CEST esteja limpo."""
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            # Escreve o cabeçalho (usando nomes consistentes)
            writer.writerow(['Descrição item', 'NCM', 'Aliq. ICMS', 'TRIBUTACAO', 'CEST'])
            for desc, values in configs.items():
                ncm = str(values.get('NCM', '')).strip()
                aliq = str(values.get('ALIQ_ICMS', '')).strip()
                trib = str(values.get('TRIBUTACAO', '')).strip() # Salva sem Ç
                cest = clean_cest(values.get('CEST', '0'))
                writer.writerow([desc, ncm, aliq, trib, cest])
        # st.sidebar.info(f"Configurações salvas em {CONFIG_FILE}") # Opcional
    except Exception as e:
        st.sidebar.error(f"Erro ao salvar configurações em '{CONFIG_FILE}': {str(e)}")

def process_planilha(df, configs):
    """Processa a planilha de auditoria comparando com as configurações."""
    # Garante a existência e limpeza inicial das colunas no DataFrame de entrada
    if 'NCM' not in df.columns:
        df['NCM'] = ''
    else:
        df['NCM'] = df['NCM'].astype(str).str.replace('\.0$', '', regex=True).str.strip()

    # Padroniza nomes de coluna de entrada (exemplo)
    df.columns = df.columns.str.strip().str.upper().str.replace('.', '', regex=False)
    col_mapping_audit = {
        'DESCRIÇÃO ITEM': 'Descrição item', 'DESCRICAO ITEM': 'Descrição item',
        'NCM': 'NCM',
        'ALIQ ICMS': 'Aliq. ICMS', 'ALIQUOTA ICMS': 'Aliq. ICMS',
        'TRIBUTAÇÃO': 'TRIBUTACAO', 'TRIBUTACAO': 'TRIBUTACAO',
        'CEST': 'CEST'
    }
    df.rename(columns=col_mapping_audit, inplace=True)

    if 'Aliq. ICMS' not in df.columns:
        df['Aliq. ICMS'] = ''
    else:
        df['Aliq. ICMS'] = df['Aliq. ICMS'].astype(str).str.strip()

    # Usa o nome padronizado sem Ç
    if 'TRIBUTACAO' not in df.columns:
        df['TRIBUTACAO'] = ''
    else:
        df['TRIBUTACAO'] = df['TRIBUTACAO'].astype(str).str.strip()

    if 'CEST' not in df.columns:
        df['CEST'] = '0'
    df['CEST'] = df['CEST'].apply(clean_cest)

    # Inicializa colunas de controle
    df['NCM Alterado'] = False
    df['Aliq. ICMS Alterado'] = False
    df['TRIBUTACAO Alterado'] = False # Nome interno sem Ç
    df['CEST Alterado'] = False
    df['ITEM CONSIDERADO'] = ''
    df['SIMILARIDADE'] = 0.0

    # Itera sobre cada linha da planilha de auditoria
    for i, row in df.iterrows():
        desc_item = str(row.get('Descrição item', '')).strip().lower()
        if not desc_item:
            continue

        ncm_item = str(row.get('NCM', '')).strip()
        aliq_item = str(row.get('Aliq. ICMS', '')).strip()
        trib_item = str(row.get('TRIBUTACAO', '')).strip() # Usa nome sem Ç
        cest_item = str(row.get('CEST', '0')).strip()

        palavras_item = get_keywords(desc_item)
        melhor_match = None
        max_score = -1
        match_type = "Nenhum"

        # 1. Procura por correspondência exata na descrição (case-insensitive)
        # Cria um mapeamento desc_lower -> desc_original para buscar
        configs_lower = {k.lower(): k for k in configs.keys()}
        if desc_item in configs_lower:
             original_desc = configs_lower[desc_item]
             melhor_match = original_desc
             max_score = 100
             match_type = "Descrição Exata"

        # 2. Se não encontrou exata, procura por palavras-chave ou NCM
        if not melhor_match:
            for desc_base, values in configs.items():
                desc_base_clean = desc_base.strip().lower()
                palavras_base = get_keywords(desc_base_clean)
                palavras_iguais = palavras_item & palavras_base
                ncm_base = str(values.get('NCM', '')).strip()

                # Considera match se tiver >= 2 palavras iguais OU NCM igual (e não vazio)
                if len(palavras_iguais) >= 2 or (ncm_base and ncm_base == ncm_item):
                    score_atual = len(palavras_iguais) * 10 + (50 if ncm_base and ncm_base == ncm_item else 0)
                    if score_atual > max_score:
                        max_score = score_atual
                        melhor_match = desc_base
                        match_type = "Palavras/NCM"

        # 3. Se ainda não encontrou, usa fuzzy matching
        if not melhor_match:
            for desc_base, values in configs.items():
                desc_base_clean = desc_base.strip().lower()
                score = fuzz.ratio(desc_item, desc_base_clean)
                # Limiar de similaridade - pode ser ajustado
                similarity_threshold = 70
                if score >= similarity_threshold and score > max_score:
                    max_score = score
                    melhor_match = desc_base
                    match_type = f"Similaridade ({score}%)"

        # Se encontrou um melhor match por qualquer método
        if melhor_match:
            valores_base = configs[melhor_match]
            ncm_base = str(valores_base.get('NCM', '')).strip()
            aliq_base = str(valores_base.get('ALIQ_ICMS', '')).strip()
            trib_base = str(valores_base.get('TRIBUTACAO', '')).strip() # Usa nome sem Ç
            cest_base = clean_cest(valores_base.get('CEST', '0'))

            # Compara e atualiza os campos, marcando as alterações
            if ncm_item != ncm_base:
                df.at[i, 'NCM'] = ncm_base
                df.at[i, 'NCM Alterado'] = True
            if aliq_item != aliq_base:
                df.at[i, 'Aliq. ICMS'] = aliq_base
                df.at[i, 'Aliq. ICMS Alterado'] = True
            if trib_item != trib_base:
                df.at[i, 'TRIBUTACAO'] = trib_base # Atualiza coluna sem Ç
                df.at[i, 'TRIBUTACAO Alterado'] = True
            if cest_item != cest_base:
                df.at[i, 'CEST'] = cest_base
                df.at[i, 'CEST Alterado'] = True

            df.at[i, 'ITEM CONSIDERADO'] = f'{match_type}: {melhor_match}'
            df.at[i, 'SIMILARIDADE'] = max_score
        else:
            df.at[i, 'ITEM CONSIDERADO'] = 'Nenhuma correspondência encontrada'
            df.at[i, 'SIMILARIDADE'] = 0

    return df

def aplicar_destaque_excel(df, filename):
    """Salva o DataFrame em Excel com destaque nas células alteradas."""
    try:
        # Cria uma cópia para não modificar o DataFrame original que pode ser usado em outro lugar
        df_copy = df.copy()

        # Renomeia a coluna de tributação para exibição no Excel (com Ç)
        if 'TRIBUTACAO' in df_copy.columns:
             df_copy.rename(columns={'TRIBUTACAO': 'TRIBUTAÇÃO'}, inplace=True)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Remove colunas de controle antes de salvar no Excel
            cols_to_drop = ['NCM Alterado', 'Aliq. ICMS Alterado', 'TRIBUTACAO Alterado', 'CEST Alterado', 'SIMILARIDADE']
            df_final_excel = df_copy.drop(columns=[col for col in cols_to_drop if col in df_copy.columns], errors='ignore')

            df_final_excel.to_excel(writer, index=False, sheet_name='Resultado Auditoria')
            workbook = writer.book
            worksheet = writer.sheets['Resultado Auditoria']

            yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

            # Encontra os índices das colunas pelo nome *no DataFrame original* (df)
            # porque precisamos dos dados das colunas 'Alterado'
            cols_orig = {col_name: idx + 1 for idx, col_name in enumerate(df.columns)}
            # Encontra os índices das colunas *no DataFrame final do Excel* (df_final_excel)
            cols_excel = {col_name: idx + 1 for idx, col_name in enumerate(df_final_excel.columns)}

            # Itera pelas linhas do DataFrame original (df) para checar alterações
            for row_idx, row_data in df.iterrows():
                excel_row = row_idx + 2 # +1 para 1-based index, +1 para pular cabeçalho

                if row_data.get('NCM Alterado', False):
                    if 'NCM' in cols_excel:
                        worksheet.cell(row=excel_row, column=cols_excel['NCM']).fill = yellow_fill
                if row_data.get('Aliq. ICMS Alterado', False):
                    if 'Aliq. ICMS' in cols_excel:
                        worksheet.cell(row=excel_row, column=cols_excel['Aliq. ICMS']).fill = yellow_fill
                if row_data.get('TRIBUTACAO Alterado', False):
                    # Destaca a coluna 'TRIBUTAÇÃO' (com Ç) no Excel
                    if 'TRIBUTAÇÃO' in cols_excel:
                        worksheet.cell(row=excel_row, column=cols_excel['TRIBUTAÇÃO']).fill = yellow_fill
                if row_data.get('CEST Alterado', False):
                    if 'CEST' in cols_excel:
                        worksheet.cell(row=excel_row, column=cols_excel['CEST']).fill = yellow_fill

        st.info(f"Arquivo Excel '{filename}' gerado com destaque.")
    except Exception as e:
        st.error(f"Erro ao gerar arquivo Excel com destaque: {str(e)}")

def export_to_pdf(df, filename):
    """Exporta o DataFrame para um arquivo PDF simples."""
    try:
        # Cria uma cópia e renomeia TRIBUTACAO para TRIBUTAÇÃO para o PDF
        df_pdf = df.copy()
        if 'TRIBUTACAO' in df_pdf.columns:
            df_pdf.rename(columns={'TRIBUTACAO': 'TRIBUTAÇÃO'}, inplace=True)

        c = canvas.Canvas(filename, pagesize=letter)
        width, height = letter
        margin = 40 # Reduzido um pouco a margem
        y_position = height - margin
        line_height = 11 # Reduzido um pouco

        # Cabeçalho do PDF
        c.setFont("Helvetica-Bold", 9)
        x_offset = margin

        # Colunas a serem exibidas no PDF (exclui controle)
        header_names = [col for col in df_pdf.columns if not col.endswith('Alterado') and col != 'SIMILARIDADE']

        # Cálculo dinâmico (e simples) das larguras das colunas
        available_width = width - 2 * margin
        # Pesos estimados (ajustar conforme necessidade)
        weights = {'Descrição item': 3, 'NCM': 1.5, 'CEST': 1, 'Aliq. ICMS': 1, 'TRIBUTAÇÃO': 1, 'ITEM CONSIDERADO': 2.5}
        total_weight = sum(weights.get(col, 1) for col in header_names)
        col_widths = {col: (weights.get(col, 1) / total_weight) * available_width for col in header_names}

        # Desenha cabeçalho da tabela
        current_x = x_offset
        for col_name in header_names:
            c.drawString(current_x + 2, y_position, col_name) # Adiciona pequeno padding
            current_x += col_widths[col_name]
        y_position -= line_height * 1.5

        # Desenha linhas da tabela
        c.setFont("Helvetica", 7) # Fonte menor para caber mais
        for _, row in df_pdf.iterrows():
            if y_position < margin + line_height:
                c.showPage()
                c.setFont("Helvetica-Bold", 9)
                y_position = height - margin
                # Redesenha cabeçalho na nova página
                current_x = x_offset
                for col_name in header_names:
                    c.drawString(current_x + 2, y_position, col_name)
                    current_x += col_widths[col_name]
                y_position -= line_height * 1.5
                c.setFont("Helvetica", 7)

            current_x = x_offset
            max_lines_per_row = 1
            row_lines_content = {} # Armazena linhas quebradas por coluna

            # Prepara o conteúdo das células, quebrando linhas se necessário
            for col_name in header_names:
                cell_value = str(row.get(col_name, ''))
                col_width_pixels = col_widths[col_name] - 4 # Largura útil com padding
                avg_char_width = 4 # Estimativa grosseira da largura do caractere
                max_chars_per_line = int(col_width_pixels / avg_char_width)

                lines = []
                if max_chars_per_line > 0:
                    import textwrap
                    lines = textwrap.wrap(cell_value, width=max_chars_per_line, break_long_words=True, max_lines=5) # Limita a 5 linhas por célula
                else:
                    lines = [cell_value[:10] + '...' if len(cell_value) > 10 else cell_value] # Fallback

                row_lines_content[col_name] = lines
                max_lines_per_row = max(max_lines_per_row, len(lines))

            # Desenha as linhas da célula atual (pode ocupar múltiplas linhas no PDF)
            initial_y = y_position
            for line_num in range(max_lines_per_row):
                current_x = x_offset
                y_line_pos = initial_y - (line_num * line_height * 0.9) # Espaçamento entre linhas dentro da célula
                for col_name in header_names:
                    lines = row_lines_content[col_name]
                    if line_num < len(lines):
                        c.drawString(current_x + 2, y_line_pos, lines[line_num])
                    current_x += col_widths[col_name]

            y_position = initial_y - (max_lines_per_row * line_height * 0.9) - (line_height * 0.3) # Espaço extra entre linhas da tabela

        c.save()
        st.info(f"Arquivo PDF '{filename}' gerado.")
    except ImportError:
        st.error("Erro ao gerar PDF: A biblioteca 'textwrap' não foi encontrada. PDF não gerado.")
    except Exception as e:
        st.error(f"Erro ao gerar arquivo PDF: {str(e)}")

# --- Interface Streamlit --- 

# Carregar configurações iniciais
configs = load_configurations()

# Informações de status na barra lateral
st.sidebar.write("### Status do Sistema")
st.sidebar.write(f"Itens na base de configurações: {len(configs)}")
if not configs:
    st.sidebar.warning("⚠️ Base de configurações vazia. Adicione uma base na Aba 1 ou verifique os arquivos 'configuracoes.csv'/'configuracoes.xlsx'.")
else:
    st.sidebar.success("✅ Base de configurações carregada.")

# Criação das abas no Streamlit
tab1, tab2, tab3 = st.tabs(["🟩 1. Adicionar/Atualizar Base", "📋 2. Ver Base de Configurações", "🔍 3. Realizar Auditoria"])

# Aba 1 - Adicionar/Atualizar Base Auditada
with tab1:
    st.header("1. Adicionar ou Atualizar Base de Configurações")
    st.markdown("Envie uma planilha Excel (`.xlsx`) auditada para adicionar novos itens ou atualizar existentes na base de configurações.")
    uploaded_base = st.file_uploader("Selecione a planilha auditada", type=['xlsx'], key='base_uploader')

    if uploaded_base:
        try:
            base_df = pd.read_excel(uploaded_base, dtype={'NCM': str, 'CEST': str})
            # Padroniza nomes de coluna do upload
            base_df.columns = base_df.columns.str.strip().str.replace('"', '', regex=False).str.replace('\n', '', regex=False).str.replace('\r', '', regex=False).str.upper()
            base_df.columns = base_df.columns.str.replace('.', '', regex=False)
            col_mapping_upload = {
                'DESCRIÇÃO ITEM': 'Descrição item', 'DESCRICAO ITEM': 'Descrição item',
                'NCM': 'NCM',
                'ALIQ ICMS': 'Aliq. ICMS', 'ALIQUOTA ICMS': 'Aliq. ICMS',
                'TRIBUTAÇÃO': 'TRIBUTACAO', 'TRIBUTACAO': 'TRIBUTACAO',
                'CEST': 'CEST'
            }
            base_df.rename(columns=col_mapping_upload, inplace=True)

            required_cols = ['Descrição item', 'NCM', 'Aliq. ICMS']
            missing_cols = [col for col in required_cols if col not in base_df.columns]
            if missing_cols:
                 st.error(f"Erro: A planilha enviada deve conter as colunas: {', '.join(required_cols)}. Colunas ausentes ou não reconhecidas: {', '.join(missing_cols)}")
            else:
                itens_adicionados = 0
                itens_atualizados = 0
                for _, row in base_df.iterrows():
                    desc = str(row['Descrição item']).strip()
                    if not desc: continue

                    ncm = str(row['NCM']).strip() if pd.notna(row['NCM']) else ''
                    aliq = str(row['Aliq. ICMS']).strip()
                    trib = str(row.get('TRIBUTACAO', aliq)).strip() # Usa nome sem Ç
                    cest = clean_cest(row.get('CEST', '0'))

                    if desc in configs:
                        itens_atualizados += 1
                    else:
                        itens_adicionados += 1

                    configs[desc] = {'NCM': ncm, 'ALIQ_ICMS': aliq, 'TRIBUTACAO': trib, 'CEST': cest}

                save_all_configurations(configs)
                st.success(f"✅ Base atualizada com sucesso! Itens adicionados: {itens_adicionados}, Itens atualizados: {itens_atualizados}. Total na base: {len(configs)}.")
                st.sidebar.write(f"Itens na base de configurações: {len(configs)}")
                # Limpa o uploader para permitir novo upload sem recarregar a página manualmente
                # st.session_state.base_uploader = None # Pode causar problemas dependendo da versão do Streamlit
                st.rerun()

        except Exception as e:
            st.error(f"Erro ao processar a planilha enviada: {str(e)}")

# Aba 2 - Visualizar Base de Configurações
with tab2:
    st.header("2. Ver Base de Configurações Salva")
    st.markdown("Visualize e pesquise os itens atualmente na base de configurações.")

    if not configs:
        st.warning("⚠️ A base de configurações está vazia.")
    else:
        search_term = st.text_input("🔎 Pesquisar por Descrição, NCM, CEST ou Tributação", key='search_base')
        show_all = st.checkbox("👁️ Mostrar toda a base", key='show_all_base')

        if configs:
            df_base = pd.DataFrame.from_dict(configs, orient='index')
            df_base = df_base.reset_index().rename(columns={'index': 'Descrição item'})
            # --- Correção da Exibição da Tributação ---
            # Garante que a coluna TRIBUTACAO (sem Ç) seja selecionada e renomeada para exibição
            display_cols_internal = ['Descrição item', 'NCM', 'CEST', 'ALIQ_ICMS', 'TRIBUTACAO']
            # Seleciona apenas colunas existentes no DataFrame
            cols_to_display = [col for col in display_cols_internal if col in df_base.columns]
            df_display = df_base[cols_to_display]
            # Renomeia TRIBUTACAO para TRIBUTAÇÃO para melhor leitura na interface
            if 'TRIBUTACAO' in df_display.columns:
                df_display = df_display.rename(columns={'TRIBUTACAO': 'TRIBUTAÇÃO'})
            # Renomeia ALIQ_ICMS para Aliq. ICMS
            if 'ALIQ_ICMS' in df_display.columns:
                df_display = df_display.rename(columns={'ALIQ_ICMS': 'Aliq. ICMS'})

        else:
            df_display = pd.DataFrame(columns=['Descrição item', 'NCM', 'CEST', 'Aliq. ICMS', 'TRIBUTAÇÃO'])

        if search_term:
            search_term_lower = search_term.lower().strip()
            # Busca em todas as colunas do df_display
            mask = df_display.apply(lambda row: any(search_term_lower in str(cell).lower() for cell in row), axis=1)
            df_filtrada = df_display[mask]
            st.dataframe(df_filtrada, use_container_width=True)
            st.caption(f"{len(df_filtrada)} itens encontrados para '{search_term}'.")
        elif show_all:
            st.dataframe(df_display, use_container_width=True)
            st.caption(f"Mostrando todos os {len(df_display)} itens da base.")
        else:
            st.info("Digite um termo de pesquisa ou marque 'Mostrar toda a base' para ver os dados.")

# Aba 3 - Realizar Auditoria
with tab3:
    st.header("3. Realizar Auditoria")
    st.markdown("Envie uma planilha Excel (`.xlsx`) para ser auditada com base nas configurações atuais.")

    if not configs:
        st.warning("⚠️ A base de configurações está vazia. Adicione uma base na Aba 1 primeiro.")
    else:
        uploaded_audit = st.file_uploader("Selecione a planilha para auditoria", type=['xlsx'], key='audit_uploader')

        if uploaded_audit:
            try:
                audit_df = pd.read_excel(uploaded_audit, dtype={'NCM': str, 'CEST': str})
                # Padroniza nomes de coluna da auditoria
                audit_df.columns = audit_df.columns.str.strip().str.replace('"', '', regex=False).str.replace('\n', '', regex=False).str.replace('\r', '', regex=False).str.upper()
                audit_df.columns = audit_df.columns.str.replace('.', '', regex=False)
                col_mapping_audit_upload = {
                    'DESCRIÇÃO ITEM': 'Descrição item', 'DESCRICAO ITEM': 'Descrição item',
                    'NCM': 'NCM',
                    'ALIQ ICMS': 'Aliq. ICMS', 'ALIQUOTA ICMS': 'Aliq. ICMS',
                    'TRIBUTAÇÃO': 'TRIBUTACAO', 'TRIBUTACAO': 'TRIBUTACAO',
                    'CEST': 'CEST'
                }
                audit_df.rename(columns=col_mapping_audit_upload, inplace=True)

                if 'Descrição item' not in audit_df.columns:
                    st.error("Erro: A planilha de auditoria deve conter a coluna 'Descrição item'.")
                else:
                    progress_bar = st.progress(0, text="Processando auditoria...")
                    # Processa a planilha (passa a barra de progresso se a função suportar)
                    # Nota: A função process_planilha atual não tem suporte a barra de progresso.
                    # Para adicionar, seria necessário passar a barra e atualizá-la dentro do loop.
                    result_df = process_planilha(audit_df.copy(), configs)
                    progress_bar.progress(100, text="Auditoria concluída!")
                    st.success("✅ Auditoria concluída com sucesso!")

                    output_excel_file = "resultado_auditoria.xlsx"
                    output_pdf_file = "resultado_auditoria.pdf"

                    # Gera Excel com destaque
                    aplicar_destaque_excel(result_df.copy(), output_excel_file)
                    # Gera PDF (versão melhorada)
                    export_to_pdf(result_df.copy(), output_pdf_file)

                    st.dataframe(result_df.head(50), use_container_width=True)
                    st.caption("Prévia das primeiras 50 linhas do resultado.")

                    col1, col2 = st.columns(2)
                    try:
                        with open(output_excel_file, 'rb') as f_excel:
                            col1.download_button(
                                label="📥 Baixar Resultado (Excel)",
                                data=f_excel,
                                file_name=output_excel_file,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key='download_excel'
                            )
                    except FileNotFoundError:
                        col1.error(f"Erro: Não foi possível encontrar {output_excel_file}")
                    except Exception as e:
                         col1.error(f"Erro ao preparar download Excel: {e}")

                    try:
                        with open(output_pdf_file, 'rb') as f_pdf:
                            col2.download_button(
                                label="📥 Baixar Resultado (PDF)",
                                data=f_pdf,
                                file_name=output_pdf_file,
                                mime="application/pdf",
                                key='download_pdf'
                            )
                    except FileNotFoundError:
                        col2.error(f"Erro: Não foi possível encontrar {output_pdf_file}")
                    except Exception as e:
                         col2.error(f"Erro ao preparar download PDF: {e}")

            except Exception as e:
                st.error(f"Erro ao processar a auditoria: {str(e)}")
                # Limpa a barra de progresso em caso de erro
                if 'progress_bar' in locals():
                    progress_bar.empty()

